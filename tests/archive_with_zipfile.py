import csv
from zipfile import ZipFile
import os
import os.path
from PyPDF2 import PdfReader
from openpyxl import load_workbook

from utils.file_extension import get_file_extension


def test_zip_files(dir_check, file_remove):
    zip_file = ZipFile('../resources/sample.zip', 'w')

    zip_file.write('docs-pytest-org-en-latest.pdf')
    zip_file.write('file_example_100.xlsx')
    zip_file.write('username.csv')

    file_names = zip_file.namelist()

    for filename in file_names:

        if get_file_extension(filename) == '.pdf':
            zip_file.extract(filename, 'tmp')
            reader = PdfReader(filename)
            number_of_pages = len(reader.pages)
            page = reader.pages[0]
            text = page.extract_text()
            print(text)
            assert "pytest Documentation" in text
            assert number_of_pages == 412
            assert os.path.getsize(filename) == 1739506

        elif get_file_extension(filename) == '.xlsx':
            zip_file.extract(filename, 'tmp')
            workbook = load_workbook(filename)
            sheet = workbook.active
            assert 'Vincenza' in sheet.cell(row=10, column=2).value
            assert os.path.getsize(filename) == 9299

        elif get_file_extension(filename) == '.csv':
            zip_file.extract(filename, 'tmp')
            with open(filename, encoding="utf8") as f:
                reader = csv.DictReader(f, delimiter=";")
                list_f = list(reader)
                assert "Johnson" in (list_f[2]['Last name'])
                assert os.path.getsize(filename) == 176
        else:
            print(filename, "n/a")

    zip_file.close()
