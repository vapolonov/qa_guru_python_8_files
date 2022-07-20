from openpyxl import load_workbook

workbook = load_workbook('resources/file_example_100.xlsx')
sheet = workbook.active
print(sheet.cell(row=3, column=3).value)  # не индексы, нумерация с 1
workbook.close()
